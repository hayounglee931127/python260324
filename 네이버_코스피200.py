import sys
import requests
from bs4 import BeautifulSoup
import time
import csv
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QPushButton, QTableWidget, QTableWidgetItem,
    QVBoxLayout, QHBoxLayout, QWidget, QProgressBar, QLabel, QMessageBox,
    QFileDialog, QTextEdit
)
from PyQt6.QtCore import QThread, pyqtSignal
import pandas as pd
from openpyxl import Workbook

class CrawlWorker(QThread):
    progress = pyqtSignal(int)
    finished = pyqtSignal(list)
    error = pyqtSignal(str)

    def run(self):
        try:
            items = self.scrape_kospi200_constituents()
            self.finished.emit(items)
        except Exception as e:
            self.error.emit(str(e))

    def scrape_kospi200_constituents(self):
        """코스피200 편입종목 전체 크롤링"""

        base_url = "https://finance.naver.com/sise/entryJongmok.naver?type=KPI200"

        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            )
        }

        all_items = []

        # 코스피200은 총 200개 종목, 페이지당 10개씩 표시되므로 20페이지
        total_pages = 20

        for page in range(1, total_pages + 1):
            url = f"{base_url}&page={page}"

            try:
                resp = requests.get(url, headers=headers, timeout=10)
                resp.raise_for_status()

                soup = BeautifulSoup(resp.text, "html.parser")

                # 편입종목 테이블 찾기
                table = soup.select_one("table.type_1")
                if not table:
                    continue

                rows = table.select("tr")

                page_items = []
                for row in rows:
                    cols = [td.get_text(strip=True) for td in row.select("td")]
                    # 헤더 행 제외 (종목별, 현재가 등이 있는 행)
                    if cols and len(cols) >= 7 and "종목별" not in cols[0]:
                        # 종목명, 현재가, 전일비, 등락률, 거래량, 거래대금, 시가총액
                        if len(cols) >= 7:
                            item = {
                                "종목명": cols[0],
                                "현재가": cols[1],
                                "전일비": cols[2],
                                "등락률": cols[3],
                                "거래량": cols[4],
                                "거래대금": cols[5],
                                "시가총액": cols[6]
                            }
                            page_items.append(item)

                all_items.extend(page_items)

                time.sleep(0.5)  # 서버 부하 방지

            except Exception as e:
                raise e

            self.progress.emit(int((page / total_pages) * 100))

        return all_items

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("코스피200 크롤러")
        self.setGeometry(100, 100, 800, 600)

        self.items = []

        # 위젯 생성
        self.start_button = QPushButton("크롤링 시작")
        self.start_button.clicked.connect(self.start_crawl)

        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)

        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(["종목명", "현재가", "전일비", "등락률", "거래량", "거래대금", "시가총액"])

        self.save_button = QPushButton("Excel 저장")
        self.save_button.clicked.connect(self.save_excel)
        self.save_button.setEnabled(False)

        self.log_text = QTextEdit()
        self.log_text.setMaximumHeight(100)

        # 레이아웃
        layout = QVBoxLayout()
        layout.addWidget(self.start_button)
        layout.addWidget(self.progress_bar)
        layout.addWidget(self.table)
        layout.addWidget(self.save_button)
        layout.addWidget(QLabel("로그:"))
        layout.addWidget(self.log_text)

        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

    def start_crawl(self):
        self.start_button.setEnabled(False)
        self.progress_bar.setValue(0)
        self.log_text.clear()
        self.log_text.append("크롤링 시작...")

        self.worker = CrawlWorker()
        self.worker.progress.connect(self.update_progress)
        self.worker.finished.connect(self.on_finished)
        self.worker.error.connect(self.on_error)
        self.worker.start()

    def update_progress(self, value):
        self.progress_bar.setValue(value)

    def on_finished(self, items):
        self.items = items
        self.table.setRowCount(len(items))
        for row, item in enumerate(items):
            self.table.setItem(row, 0, QTableWidgetItem(item["종목명"]))
            self.table.setItem(row, 1, QTableWidgetItem(item["현재가"]))
            self.table.setItem(row, 2, QTableWidgetItem(item["전일비"]))
            self.table.setItem(row, 3, QTableWidgetItem(item["등락률"]))
            self.table.setItem(row, 4, QTableWidgetItem(item["거래량"]))
            self.table.setItem(row, 5, QTableWidgetItem(item["거래대금"]))
            self.table.setItem(row, 6, QTableWidgetItem(item["시가총액"]))

        self.save_button.setEnabled(True)
        self.start_button.setEnabled(True)
        self.log_text.append(f"크롤링 완료. 총 {len(items)}개 종목 수집.")

    def on_error(self, error_msg):
        QMessageBox.critical(self, "오류", f"크롤링 중 오류 발생: {error_msg}")
        self.start_button.setEnabled(True)
        self.log_text.append(f"오류: {error_msg}")

    def save_excel(self):
        if not self.items:
            QMessageBox.warning(self, "경고", "저장할 데이터가 없습니다.")
            return

        file_path, _ = QFileDialog.getSaveFileName(self, "Excel 파일 저장", "kospi200.xlsx", "Excel Files (*.xlsx)")
        if file_path:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "코스피200"

                # 헤더 추가
                headers = ["종목명", "현재가", "전일비", "등락률", "거래량", "거래대금", "시가총액"]
                for col_num, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_num, value=header)

                # 데이터 추가
                for row_num, item in enumerate(self.items, 2):
                    ws.cell(row=row_num, column=1, value=item["종목명"])
                    ws.cell(row=row_num, column=2, value=item["현재가"])
                    ws.cell(row=row_num, column=3, value=item["전일비"])
                    ws.cell(row=row_num, column=4, value=item["등락률"])
                    ws.cell(row=row_num, column=5, value=item["거래량"])
                    ws.cell(row=row_num, column=6, value=item["거래대금"])
                    ws.cell(row=row_num, column=7, value=item["시가총액"])

                wb.save(file_path)
                QMessageBox.information(self, "성공", f"Excel 파일로 저장되었습니다: {file_path}")
                self.log_text.append(f"Excel 저장 완료: {file_path}")
            except Exception as e:
                QMessageBox.critical(self, "오류", f"Excel 저장 실패: {e}")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())