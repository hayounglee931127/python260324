import openpyxl
import random

# 전자제품 이름 리스트
product_names = [
    "스마트폰", "노트북", "태블릿", "헤드폰", "키보드", "마우스", "모니터", "프린터", "카메라", "스피커",
    "충전기", "케이블", "하드드라이브", "SSD", "메모리카드", "블루투스 이어폰", "웹캠", "마이크", "프로젝터", "스마트워치"
]

# 워크북 생성
wb = openpyxl.Workbook()
ws = wb.active

# 헤더 작성
ws['A1'] = '제품ID'
ws['B1'] = '제품명'
ws['C1'] = '가격'
ws['D1'] = '수량'

# 100개의 데이터 생성 및 작성
for i in range(2, 102):  # 2행부터 101행까지 (헤더 제외)
    product_id = i - 1  # 제품ID는 1부터 시작
    product_name = random.choice(product_names)
    price = random.randint(10000, 1000000)
    quantity = random.randint(1, 100)

    ws[f'A{i}'] = product_id
    ws[f'B{i}'] = product_name
    ws[f'C{i}'] = price
    ws[f'D{i}'] = quantity

# 파일 저장
wb.save('ProductList.xlsx')
print("ProductList.xlsx 파일이 생성되었습니다.")